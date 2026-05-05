"""등급컷 입력 양식 엑셀 생성 스크립트.
실행: python _make_template.py
출력: cuts_template.xlsx
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

wb = Workbook()

# ── 1. README 시트 ────────────────────────────────
readme = wb.active
readme.title = 'README'

readme_rows = [
    ('등급컷 아카이브 입력 양식', '', ''),
    ('', '', ''),
    ('이 파일은 grade-cut 페이지의 cuts.json으로 변환되는 입력 양식입니다.', '', ''),
    ('"data" 시트의 빈 행을 채워 넣고 저장한 뒤 채훈에게 보내주세요.', '', ''),
    ('', '', ''),
    ('━━ 시트 구조 ━━', '', ''),
    ('', '', ''),
    ('시트', '용도', '비고'),
    ('README', '이 안내', '편집 X'),
    ('data', '실제 데이터 입력', '한 행 = (시험 × 과목)'),
    ('lookup', '값 참조용 드롭다운', '편집 X'),
    ('', '', ''),
    ('━━ data 시트 컬럼 ━━', '', ''),
    ('', '', ''),
    ('컬럼', '의미', '예시'),
    ('year', '학년도 (수능 기준)', '2026'),
    ('type', '시험 종류', 'suneung / pyeongga / gyoyukcheong'),
    ('exam_name', '시험 정식 명칭', '2026학년도 대학수학능력시험'),
    ('exam_date', '시험 시행일 (YYYY-MM-DD)', '2025-11-13'),
    ('source', '출처', '한국교육과정평가원 / 교육청 등'),
    ('verified', '공식 발표 여부', 'TRUE 또는 FALSE'),
    ('subject', '과목', 'english / korean-history (절대평가만 우선)'),
    ('g1 ~ g8', '1~8등급 컷 원점수', '90, 80, 70, 60, 50, 40, 30, 20'),
    ('pct1 ~ pct9', '1~9등급의 등급별 비율 (%) (누적 X)', '3.11, 14.35, 26.30, ...'),
    ('', '', ''),
    ('━━ 입력 규칙 ━━', '', ''),
    ('', '', ''),
    ('1. 같은 시험 여러 과목은 별개 행으로 입력 (예: 영어 1행, 한국사 1행)', '', ''),
    ('2. type 값은 lookup 시트의 옵션 중 하나만 사용', '', ''),
    ('3. subject 값은 lookup 시트의 옵션 중 하나만 사용', '', ''),
    ('4. pct는 누적이 아니라 등급별 비율 (합계 ≈ 100%)', '', ''),
    ('5. 컷 점수는 1등급(가장 높음) → 8등급 순서로 내림차순', '', ''),
    ('6. 모르는 값이 있으면 빈칸으로 둬도 OK', '', ''),
    ('', '', ''),
    ('━━ 우선순위 (영어/한국사부터) ━━', '', ''),
    ('', '', ''),
    ('1순위: 22~26 수능 영어/한국사 (5년 × 2과목 = 10행)', '', ''),
    ('2순위: 22~26 평가원 6·9월 모의평가 영어/한국사 (5년 × 2회 × 2과목 = 20행)', '', ''),
    ('3순위: 교육청 모의고사 영어/한국사', '', ''),
    ('', '', ''),
    ('국어/수학/탐구는 다른 양식으로 추후 진행', '', ''),
]

for row in readme_rows:
    readme.append(row)

readme.column_dimensions['A'].width = 38
readme.column_dimensions['B'].width = 40
readme.column_dimensions['C'].width = 50
readme['A1'].font = Font(bold=True, size=16, color='667EEA')
for cell in (readme['A6'], readme['A13'], readme['A26'], readme['A35']):
    cell.font = Font(bold=True, size=12, color='764BA2')
for cell in (readme['A8'], readme['A15']):
    cell.font = Font(bold=True)
    readme[cell.coordinate.replace('A','B')].font = Font(bold=True)
    readme[cell.coordinate.replace('A','C')].font = Font(bold=True)


# ── 2. lookup 시트 (드롭다운 옵션 정의) ──────────
lookup = wb.create_sheet('lookup')
lookup.append(('type 옵션', 'subject 옵션'))
lookup['A1'].font = Font(bold=True)
lookup['B1'].font = Font(bold=True)

types = ['suneung', 'pyeongga', 'gyoyukcheong']
subjects = ['english', 'korean-history', 'korean', 'math']
for i in range(max(len(types), len(subjects))):
    t = types[i] if i < len(types) else None
    s = subjects[i] if i < len(subjects) else None
    lookup.append((t, s))

# 라벨 안내
lookup['D1'] = 'type 라벨'
lookup['E1'] = ''
lookup['D1'].font = Font(bold=True)
type_labels = [
    ('suneung',     '수능'),
    ('pyeongga',    '평가원 모의평가 (6월/9월)'),
    ('gyoyukcheong','교육청 모의고사 (3·4·7·10월 등)'),
]
for i, (k, v) in enumerate(type_labels, start=2):
    lookup.cell(row=i, column=4, value=k)
    lookup.cell(row=i, column=5, value=v)

lookup.column_dimensions['A'].width = 16
lookup.column_dimensions['B'].width = 20
lookup.column_dimensions['D'].width = 16
lookup.column_dimensions['E'].width = 30


# ── 3. data 시트 (메인 입력) ──────────────────────
data = wb.create_sheet('data', index=0)  # 첫 시트로
wb.move_sheet('README', offset=2)         # README는 뒤로

headers = [
    'year', 'type', 'exam_name', 'exam_date', 'source', 'verified',
    'subject',
    'g1','g2','g3','g4','g5','g6','g7','g8',
    'pct1','pct2','pct3','pct4','pct5','pct6','pct7','pct8','pct9',
    'note',
]
data.append(headers)

# 헤더 스타일
header_fill = PatternFill('solid', fgColor='667EEA')
header_font = Font(bold=True, color='FFFFFF')
center = Alignment(horizontal='center', vertical='center')
thin = Side(border_style='thin', color='BBBBBB')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for col_idx, _ in enumerate(headers, start=1):
    cell = data.cell(row=1, column=col_idx)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center
    cell.border = border

# 컬럼 너비
widths = {
    'year':6, 'type':14, 'exam_name':36, 'exam_date':12, 'source':22, 'verified':9,
    'subject':16,
    'g1':5,'g2':5,'g3':5,'g4':5,'g5':5,'g6':5,'g7':5,'g8':5,
    'pct1':7,'pct2':7,'pct3':7,'pct4':7,'pct5':7,'pct6':7,'pct7':7,'pct8':7,'pct9':7,
    'note':24,
}
for i, h in enumerate(headers, start=1):
    data.column_dimensions[get_column_letter(i)].width = widths.get(h, 10)

# 샘플 행 (1줄, 채훈이 보낸 2026 영어)
sample = [
    2026, 'suneung', '2026학년도 대학수학능력시험', '2025-11-13',
    '한국교육과정평가원', True, 'english',
    90, 80, 70, 60, 50, 40, 30, 20,
    3.11, 14.35, 26.30, 24.53, 13.28, 8.02, 5.64, 3.65, 1.13,
    '샘플 — 변환 시 자동 인식'
]
data.append(sample)
for col_idx in range(1, len(headers)+1):
    data.cell(row=2, column=col_idx).border = border

# 빈 행 200개 추가 (충분한 입력 공간)
for _ in range(200):
    data.append([None]*len(headers))

# ── 데이터 유효성 검증 ────────────────
last_row = data.max_row

# type 드롭다운 (B열, 2~끝)
dv_type = DataValidation(type='list', formula1='=lookup!$A$2:$A$4', allow_blank=True)
dv_type.error = 'type은 suneung / pyeongga / gyoyukcheong 중 하나여야 합니다.'
dv_type.errorTitle = '잘못된 값'
data.add_data_validation(dv_type)
dv_type.add(f'B2:B{last_row}')

# subject 드롭다운 (G열)
dv_subject = DataValidation(type='list', formula1='=lookup!$B$2:$B$5', allow_blank=True)
dv_subject.error = 'subject는 english / korean-history / korean / math 중 하나여야 합니다.'
dv_subject.errorTitle = '잘못된 값'
data.add_data_validation(dv_subject)
dv_subject.add(f'G2:G{last_row}')

# verified 체크박스 (F열)
dv_verified = DataValidation(type='list', formula1='"TRUE,FALSE"', allow_blank=True)
data.add_data_validation(dv_verified)
dv_verified.add(f'F2:F{last_row}')

# year 숫자 (A열)
dv_year = DataValidation(type='whole', operator='between', formula1=2020, formula2=2030, allow_blank=True)
dv_year.error = 'year는 2020~2030 사이 정수.'
dv_year.errorTitle = '잘못된 값'
data.add_data_validation(dv_year)
dv_year.add(f'A2:A{last_row}')

# 첫 행 고정
data.freeze_panes = 'A2'

# ── 저장 ─────────────────────────────────────────
out = 'cuts_template.xlsx'
wb.save(out)
print(f'wrote {out}  ({last_row-1} input rows ready, 1 sample pre-filled)')
